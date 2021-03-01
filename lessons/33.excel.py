import openpyxl
import os
import pprintpp as pp
import pandas as pd

os.chdir('ldx\\EditarIES')

# wb = openpyxl.load_workbook('Correlações IES-Ploomes.xlsx')
# # ws = wb.sheetnames 		lista todas as planilhas
# ws = wb.worksheets[0]
# # tbl = ws.tables.items()[0][0]			pega a primeira tabela da planilha

# relacionados = {}
# for i in range(2,len(ws['A'])):
# 	Cod = ws.cell(row=i, column=1).value
# 	IES = ws.cell(row=i, column=2).value
# 	if IES != '-':
# 		relacionados[IES] = Cod

# pp.pprint(relacionados)
# print(len(relacionados))

xlsx = pd.ExcelFile('Correlações IES-Ploomes.xlsx')
df = xlsx.parse(xlsx.sheet_names[0])
correl = df.set_index('IES').to_dict()['Código']

pp.pprint(correl)
print(len(correl))